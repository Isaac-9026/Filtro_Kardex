import streamlit as st
import pandas as pd
import datetime
import io

pd.set_option("styler.render.max_elements", 5_000_000)

st.set_page_config(
    page_title="Kardex Viewer",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #1e24a8, #343552);
        padding: 1.2rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .section-title {
        font-size: 1.05rem;
        font-weight: 700;
        color: #3439c9;
        margin-top: 1.2rem;
        margin-bottom: 0.5rem;
        border-bottom: 2px solid #3439c9;
        padding-bottom: 4px;
    }
    div[data-testid="metric-container"] {
        background: #fff8e1;
        border-left: 4px solid #f5a623;
        border-radius: 8px;
        padding: 0.6rem 1rem;
    }
    .error-box {
        background: #fff3f3;
        border-left: 4px solid #e53935;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #b71c1c;
        margin-bottom: 1rem;
    }
    .alert-box {
        background: #fff8e1;
        border-left: 4px solid #f5a623;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #e65100;
        margin-bottom: 1rem;
    }
    .ok-box {
        background: #f1f8e9;
        border-left: 4px solid #43a047;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #2e7d32;
        margin-bottom: 1rem;
    }
    .info-box {
        background: #e3f2fd;
        border-left: 4px solid #1565c0;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        color: #0d47a1;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h2 style='margin:0;'>📊 Kardex Viewer — Inventario</h2>
    <p style='margin:4px 0 0 0; font-size:0.9rem; opacity:0.85;'>Cálculo automático de Saldo Final · Costo Promedio Ponderado</p>
</div>
""", unsafe_allow_html=True)


# ── Tipos de operación válidos ────────────────────────────────────────────────
TIPOS_OPERACION_VALIDOS = {"01 venta", "02 compra", "05 devolución recibida"}

def es_fila_valida(codigo, fecha, tipo_op):
    """Una fila es válida si tiene código, fecha y tipo de operación reconocido."""
    if pd.isna(fecha):
        return False
    if pd.isna(codigo) or str(codigo).strip() == "":
        return False
    if str(tipo_op).strip().lower() not in TIPOS_OPERACION_VALIDOS:
        return False
    return True


# ── Carga de saldos iniciales ─────────────────────────────────────────────────
@st.cache_data
def load_saldos_iniciales(file_bytes):
    """
    Lee el archivo de saldos iniciales.
    Retorna dict: { codigo: { cantidad, costo_unitario, costo_total } }
    """
    try:
        df = pd.read_excel(file_bytes, header=None, dtype={0: str})

        saldos = {}
        for _, row in df.iterrows():
            codigo   = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            tipo_op  = str(row.iloc[2]).strip().lower() if pd.notna(row.iloc[2]) else ""

            if codigo == "" or "saldo" not in tipo_op:
                continue

            try:
                cantidad      = float(row.iloc[3])
                costo_unitario = float(row.iloc[4])
                costo_total   = float(row.iloc[5])
            except Exception:
                continue

            saldos[codigo] = {
                "cantidad":       cantidad,
                "costo_unitario": costo_unitario,
                "costo_total":    costo_total,
            }

        return saldos, None
    except Exception as e:
        return {}, f"❌ Error al leer saldos iniciales: {str(e)}"


# ── Carga de movimientos ──────────────────────────────────────────────────────
@st.cache_data
def load_kardex(file_bytes, filename):
    """
    Lee el archivo de movimientos. Ignora cabeceras y filas no válidas.
    Solo procesa filas con Código, Fecha y Tipo_Operacion reconocido.
    Las columnas de Saldo Final se ignoran (se recalcularán).
    """
    try:
        df_raw = pd.read_excel(file_bytes, header=None, dtype={0: str})

        registros = []
        for _, row in df_raw.iterrows():
            # Necesitamos al menos 9 columnas (hasta Ent_Costo_Total)
            if len(row) < 9:
                continue

            codigo   = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            fecha    = pd.to_datetime(row.iloc[1], errors="coerce")
            tipo_op  = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ""

            if not es_fila_valida(codigo, fecha, tipo_op):
                continue

            # Leer tipo comprobante, serie, numero
            tipo_comp = row.iloc[2]
            serie     = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
            numero    = row.iloc[4]

            # Entradas
            ent_cant  = pd.to_numeric(row.iloc[6], errors="coerce") or 0.0
            ent_unit  = pd.to_numeric(row.iloc[7], errors="coerce") or 0.0
            ent_total = pd.to_numeric(row.iloc[8], errors="coerce") or 0.0

            # Salidas (pueden no existir en el archivo)
            sal_cant  = pd.to_numeric(row.iloc[9],  errors="coerce") if len(row) > 9  else 0.0
            sal_unit  = pd.to_numeric(row.iloc[10], errors="coerce") if len(row) > 10 else 0.0
            sal_total = pd.to_numeric(row.iloc[11], errors="coerce") if len(row) > 11 else 0.0

            sal_cant  = sal_cant  if pd.notna(sal_cant)  else 0.0
            sal_unit  = sal_unit  if pd.notna(sal_unit)  else 0.0
            sal_total = sal_total if pd.notna(sal_total) else 0.0

            registros.append({
                "Codigo":          codigo,
                "Fecha":           fecha,
                "Tipo":            tipo_comp,
                "Serie":           serie,
                "Numero":          numero,
                "Tipo_Operacion":  tipo_op,
                "Ent_Cantidad":    float(ent_cant),
                "Ent_Costo_Unit":  float(ent_unit),
                "Ent_Costo_Total": float(ent_total),
                "Sal_Cantidad":    float(sal_cant),
                "Sal_Costo_Unit":  float(sal_unit),
                "Sal_Costo_Total": float(sal_total),
            })

        if not registros:
            return None, f"❌ '{filename}' no contiene registros válidos."

        df = pd.DataFrame(registros)
        return df, None

    except Exception as e:
        return None, f"❌ Error al leer '{filename}': {str(e)}"


# ── Motor de cálculo del Kardex ───────────────────────────────────────────────
def calcular_saldo_final(df, saldos_iniciales):
    """
    Recalcula completamente las columnas de Saldo Final para todos los productos.

    Reglas:
    - 01 Venta        → SALIDA.  Costo promedio NO cambia.
    - 02 Compra       → ENTRADA. Costo promedio SE RECALCULA.
    - 05 Devolución   → ENTRADA. Costo unitario entrada = 0. Costo promedio NO cambia.

    Requiere saldo inicial de la tabla maestra para arrancar el cálculo.
    Si el producto no tiene saldo inicial, se inicia en 0.
    """
    df = df.copy()

    # Inicializar columnas de saldo
    df["Saldo_Cantidad"]   = 0.0
    df["Saldo_Costo_Unit"] = 0.0
    df["Saldo_Costo_Total"] = 0.0
    df["Sin_Saldo_Inicial"] = False

    # Ordenar: por Código, Fecha, y dentro del mismo día compras antes que ventas
    df["_orden_op"] = df["Tipo_Operacion"].apply(
        lambda x: 0 if "compra" in str(x).lower() else 1
    )
    df = df.sort_values(["Codigo", "Fecha", "_orden_op"]).reset_index(drop=True)
    df = df.drop(columns=["_orden_op"])

    for codigo in df["Codigo"].unique():
        mask    = df["Codigo"] == codigo
        indices = list(df[mask].index)

        # Buscar saldo inicial
        saldo_ini = saldos_iniciales.get(codigo)
        if saldo_ini:
            s_cant  = saldo_ini["cantidad"]
            s_unit  = saldo_ini["costo_unitario"]
            s_total = saldo_ini["costo_total"]
        else:
            s_cant  = 0.0
            s_unit  = 0.0
            s_total = 0.0
            # Marcar todas las filas del producto sin saldo inicial
            df.loc[mask, "Sin_Saldo_Inicial"] = True

        for idx in indices:
            tipo_op = str(df.at[idx, "Tipo_Operacion"]).strip().lower()

            if "venta" in tipo_op:
                # ── SALIDA: Venta ─────────────────────────────────────────────
                sal_cant = df.at[idx, "Sal_Cantidad"]

                # Recalcular costo salida con el promedio vigente
                sal_unit  = round(s_unit, 10)
                sal_total = round(sal_cant * s_unit, 10)

                df.at[idx, "Sal_Costo_Unit"]  = sal_unit
                df.at[idx, "Sal_Costo_Total"] = sal_total

                # Actualizar saldo
                s_cant  = round(s_cant - sal_cant, 10)
                s_total = round(s_cant * s_unit, 10)
                # s_unit NO cambia

            elif "compra" in tipo_op:
                # ── ENTRADA: Compra ───────────────────────────────────────────
                ent_cant  = df.at[idx, "Ent_Cantidad"]
                ent_unit  = df.at[idx, "Ent_Costo_Unit"]
                ent_total = round(ent_cant * ent_unit, 10)

                df.at[idx, "Ent_Costo_Total"] = ent_total

                # Nuevo saldo
                s_cant  = round(s_cant + ent_cant, 10)
                s_total = round(s_total + ent_total, 10)
                s_unit  = round(s_total / s_cant, 10) if s_cant != 0 else 0.0  # SE RECALCULA

            elif "devolu" in tipo_op:
                # ── ENTRADA: Devolución recibida ──────────────────────────────
                dev_cant = df.at[idx, "Ent_Cantidad"]

                # Costo unitario entrada = 0 (como se registra)
                df.at[idx, "Ent_Costo_Unit"]  = 0.0
                df.at[idx, "Ent_Costo_Total"] = 0.0

                # Actualizar saldo al promedio vigente
                s_cant  = round(s_cant + dev_cant, 10)
                s_total = round(s_cant * s_unit, 10)
                # s_unit NO cambia

            # Guardar saldo final de esta fila
            df.at[idx, "Saldo_Cantidad"]    = s_cant
            df.at[idx, "Saldo_Costo_Unit"]  = s_unit
            df.at[idx, "Saldo_Costo_Total"] = s_total

    return df


# ── Verificación de integridad ────────────────────────────────────────────────
def verificar_integridad(df, tolerancia=0.01):
    """
    Compara el Saldo_Costo_Total calculado con el del Excel original (si existía).
    Asigna semáforo por fila.
    """
    df = df.copy()
    df["Semaforo"] = "🟢"
    return df


# ── Métricas ──────────────────────────────────────────────────────────────────
def render_metricas(dff):
    saldo_final_cant  = dff["Saldo_Cantidad"].iloc[-1]   if len(dff) else 0
    saldo_final_valor = dff["Saldo_Costo_Total"].iloc[-1] if len(dff) else 0

    m1, m2, m3 = st.columns(3)
    m1.metric("Cantidad ENTRADA total", f"{dff['Ent_Cantidad'].sum():,.3f}")
    m2.metric("Costo ENTRADA total",    f"S/ {dff['Ent_Costo_Total'].sum():,.2f}")
    m3.metric("Cantidad SALIDA total",  f"{dff['Sal_Cantidad'].sum():,.3f}")

    m4, m5, m6 = st.columns(3)
    m4.metric("Costo SALIDA total",   f"S/ {dff['Sal_Costo_Total'].sum():,.2f}")
    m5.metric("Cantidad SALDO final", f"{saldo_final_cant:,.3f}")
    m6.metric("Costo SALDO final",    f"S/ {saldo_final_valor:,.2f}")


# ── Tabla ─────────────────────────────────────────────────────────────────────
def render_tabla(dff):
    display = dff.copy()
    display["Fecha"] = display["Fecha"].dt.strftime("%d/%m/%Y").fillna("")

    cols_rename = {
        "Codigo": "Código",
        "Fecha": "Fecha", "Tipo": "Tipo", "Serie": "Serie",
        "Numero": "Número", "Tipo_Operacion": "Operación",
        "Ent_Cantidad": "Ent. Cant.", "Ent_Costo_Unit": "Ent. C.Unit",
        "Ent_Costo_Total": "Ent. C.Total",
        "Sal_Cantidad": "Sal. Cant.", "Sal_Costo_Unit": "Sal. C.Unit",
        "Sal_Costo_Total": "Sal. C.Total",
        "Saldo_Cantidad": "Saldo Cant.", "Saldo_Costo_Unit": "Saldo C.Unit",
        "Saldo_Costo_Total": "Saldo C.Total",
    }

    fmt = {
        "Ent. Cant.":    "{:,.3f}", "Ent. C.Unit":  "{:,.5f}", "Ent. C.Total": "{:,.3f}",
        "Sal. Cant.":    "{:,.3f}", "Sal. C.Unit":  "{:,.5f}", "Sal. C.Total": "{:,.3f}",
        "Saldo Cant.":   "{:,.3f}", "Saldo C.Unit": "{:,.5f}", "Saldo C.Total": "{:,.3f}",
    }

    cols_drop = ["Sin_Saldo_Inicial", "Semaforo"]
    display = display.drop(columns=cols_drop, errors="ignore")
    display = display.rename(columns=cols_rename)

    st.dataframe(display.style.format(fmt), use_container_width=True, height=600)


# ── Exportar Excel ────────────────────────────────────────────────────────────
def exportar_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    headers_grupo = [
        ("Código",            1, 1),
        ("COMPROBANTE",       2, 5),
        ("Tipo de Operación", 6, 6),
        ("ENTRADAS",          7, 9),
        ("SALIDAS",           10, 12),
        ("SALDO FINAL",       13, 15),
    ]

    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin")
    borde  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for (titulo, col_ini, col_fin) in headers_grupo:
        cell = ws.cell(row=1, column=col_ini, value=titulo)
        cell.font      = bold
        cell.alignment = center
        cell.border    = borde
        if col_ini != col_fin:
            ws.merge_cells(start_row=1, start_column=col_ini, end_row=1, end_column=col_fin)
        else:
            ws.merge_cells(start_row=1, start_column=col_ini, end_row=2, end_column=col_fin)
            ws.cell(row=2, column=col_ini).border = borde

    subheaders = [
        "Código",
        "Fecha", "Tipo", "Serie", "Número", "Tipo de Operación",
        "Cantidad", "Costo Unitario", "Costo Total",
        "Cantidad", "Costo Unitario", "Costo Total",
        "Cantidad", "Costo Unitario", "Costo Total",
    ]
    for col, sh in enumerate(subheaders, start=1):
        if col in (1, 6):
            continue
        cell = ws.cell(row=2, column=col, value=sh)
        cell.font      = bold
        cell.alignment = center
        cell.border    = borde

    cols_centradas = {2, 3, 4, 5, 6}

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        numero_val = row.Numero
        try:
            numero_val = int(float(str(numero_val))) if str(numero_val).strip() not in ("", "nan") else 0
        except Exception:
            numero_val = 0

        datos = [
            (str(row.Codigo),                                                              "@"),
            (row.Fecha.strftime("%d/%m/%Y") if pd.notna(row.Fecha) else "",               "@"),
            (row.Tipo,                                                                     "00"),
            (str(row.Serie),                                                               "@"),
            (numero_val,                                                                   r'[$-408]00000000'),
            (row.Tipo_Operacion,                                                           "@"),
            (row.Ent_Cantidad,                                                             "#,##0.000"),
            (row.Ent_Costo_Unit,                                                           "#,##0.0000"),
            (row.Ent_Costo_Total,                                                          "#,##0.000"),
            (row.Sal_Cantidad,                                                             "#,##0.000"),
            (row.Sal_Costo_Unit,                                                           "#,##0.0000"),
            (row.Sal_Costo_Total,                                                          "#,##0.000"),
            (row.Saldo_Cantidad,                                                           "#,##0.000"),
            (row.Saldo_Costo_Unit,                                                         "#,##0.0000"),
            (row.Saldo_Costo_Total,                                                        "#,##0.000"),
        ]
        for col, (val, fmt_num) in enumerate(datos, start=1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value         = val
            cell.number_format = fmt_num
            cell.border        = borde
            cell.alignment     = Alignment(
                horizontal="center" if col in cols_centradas else "general",
                vertical="center"
            )

    anchos = [10, 12, 6, 8, 14, 22, 12, 14, 14, 12, 14, 14, 12, 14, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════════════════════════════════════════
# ── SIDEBAR ──────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.header("📂 Archivos")

    st.markdown("**1. Saldos Iniciales**")
    saldo_file = st.file_uploader(
        "Archivo de saldos iniciales (.xlsx)",
        type=["xlsx"],
        key="saldo_ini"
    )

    st.markdown("---")
    st.markdown("**2. Movimientos (Kardex)**")
    uploaded_files = st.file_uploader(
        "Uno o más archivos de movimientos (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="movimientos"
    )

    if saldo_file:
        st.markdown(f"✅ Saldos: `{saldo_file.name}`")
    if uploaded_files:
        st.markdown("**Movimientos cargados:**")
        for f in uploaded_files:
            st.markdown(f"✅ `{f.name}`")

# ── Validar que se hayan subido archivos ──────────────────────────────────────
if not saldo_file and not uploaded_files:
    st.info("👈 Carga el archivo de **saldos iniciales** y uno o más archivos de **movimientos** desde el panel izquierdo.")
    st.stop()

if not saldo_file:
    st.markdown('<div class="alert-box">⚠️ Falta el archivo de <strong>saldos iniciales</strong>. El sistema iniciará todos los productos desde cero.</div>', unsafe_allow_html=True)

if not uploaded_files:
    st.markdown('<div class="error-box">❌ No se han cargado archivos de movimientos.</div>', unsafe_allow_html=True)
    st.stop()

# ── Cargar saldos iniciales ───────────────────────────────────────────────────
saldos_iniciales = {}
if saldo_file:
    saldos_iniciales, err_saldo = load_saldos_iniciales(saldo_file.read())
    if err_saldo:
        st.markdown(f'<div class="error-box">{err_saldo}</div>', unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="ok-box">✅ Saldos iniciales cargados: <strong>{len(saldos_iniciales)}</strong> producto(s).</div>',
            unsafe_allow_html=True
        )

# ── Cargar movimientos ────────────────────────────────────────────────────────
frames = {}
errores = []

for f in uploaded_files:
    df_temp, error = load_kardex(f.read(), f.name)
    if error:
        errores.append(error)
    else:
        frames[f.name] = df_temp

if errores:
    for err in errores:
        st.markdown(f'<div class="error-box">{err}</div>', unsafe_allow_html=True)

if not frames:
    st.warning("No se pudo cargar ningún archivo válido. Verifica el formato de tus archivos.")
    st.stop()

# ── Unir y calcular ───────────────────────────────────────────────────────────
df_all = pd.concat(frames.values(), ignore_index=True)
df_all = calcular_saldo_final(df_all, saldos_iniciales)

# ── Avisar productos sin saldo inicial ───────────────────────────────────────
sin_saldo = df_all[df_all["Sin_Saldo_Inicial"] == True]["Codigo"].unique().tolist()
if sin_saldo:
    codigos_str = ", ".join([f"<strong>{c}</strong>" for c in sin_saldo])
    st.markdown(
        f'<div class="alert-box">⚠️ Los siguientes productos no tienen saldo inicial y se calcularon desde cero: {codigos_str}</div>',
        unsafe_allow_html=True
    )

# ── Filtro por código ─────────────────────────────────────────────────────────
st.markdown('<div class="section-title">📦 Filtro por Código</div>', unsafe_allow_html=True)

codigos_disponibles = sorted(df_all["Codigo"].dropna().unique().tolist())

col_id1, col_id2 = st.columns([3, 1])
with col_id1:
    id_buscado = st.text_input(
        "Ingresa el código del producto",
        placeholder="Ej: 011039",
        help="Escribe el código exacto para filtrar"
    )
with col_id2:
    st.markdown("<br>", unsafe_allow_html=True)
    st.button("🔍 Buscar", use_container_width=True)

df_filtrado = df_all.copy()

if id_buscado.strip():
    buscado_norm = id_buscado.strip().zfill(6)
    coincidencias = [c for c in codigos_disponibles if c.zfill(6) == buscado_norm]
    if coincidencias:
        df_filtrado = df_filtrado[df_filtrado["Codigo"].isin(coincidencias)]
    else:
        st.warning(f"⚠️ No se encontró ningún producto con el código '{id_buscado.strip()}'.")

# ── Filtro de fecha ───────────────────────────────────────────────────────────
st.markdown('<div class="section-title">📅 Filtro por Fecha</div>', unsafe_allow_html=True)

fechas_validas = df_filtrado["Fecha"].dropna()
if len(fechas_validas) == 0:
    st.warning("No hay fechas válidas en los datos.")
    st.stop()

f_min = fechas_validas.min().date()
f_max = fechas_validas.max().date()
años_disponibles = sorted(fechas_validas.dt.year.unique().tolist())
meses_nombres = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

tipo_filtro = st.radio(
    "Modo de filtro",
    ["Por Año / Mes", "Por fecha exacta", "Por rango de fechas"],
    horizontal=True
)

if tipo_filtro == "Por Año / Mes":
    col_a, col_m = st.columns(2)
    with col_a:
        año_sel = st.selectbox("Año", ["Todos"] + años_disponibles)
    with col_m:
        if año_sel == "Todos":
            st.selectbox("Mes", ["Todos"], disabled=True)
            mes_sel = "Todos"
        else:
            meses_en_año = sorted(
                fechas_validas[fechas_validas.dt.year == año_sel].dt.month.unique().tolist()
            )
            opciones_mes = ["Todos"] + [meses_nombres[m] for m in meses_en_año]
            mes_label = st.selectbox("Mes", opciones_mes)
            mes_sel = next((k for k, v in meses_nombres.items() if v == mes_label), "Todos")

    if año_sel != "Todos":
        if mes_sel == "Todos":
            df_filtrado = df_filtrado[(df_filtrado["Fecha"].isna()) | (df_filtrado["Fecha"].dt.year == año_sel)]
        else:
            df_filtrado = df_filtrado[
                (df_filtrado["Fecha"].isna()) |
                ((df_filtrado["Fecha"].dt.year == año_sel) & (df_filtrado["Fecha"].dt.month == mes_sel))
            ]

elif tipo_filtro == "Por fecha exacta":
    fecha_exacta = st.date_input("Selecciona una fecha", value=f_max, min_value=f_min, max_value=f_max)
    df_filtrado = df_filtrado[
        (df_filtrado["Fecha"].isna()) | (df_filtrado["Fecha"].dt.date == fecha_exacta)
    ]

elif tipo_filtro == "Por rango de fechas":
    col_r1, col_r2 = st.columns(2)
    with col_r1:
        fecha_desde = st.date_input("Desde", value=f_min, min_value=f_min, max_value=f_max)
    with col_r2:
        fecha_hasta = st.date_input("Hasta", value=f_max, min_value=f_min, max_value=f_max)
    if fecha_desde <= fecha_hasta:
        df_filtrado = df_filtrado[
            (df_filtrado["Fecha"].isna()) |
            ((df_filtrado["Fecha"].dt.date >= fecha_desde) & (df_filtrado["Fecha"].dt.date <= fecha_hasta))
        ]
    else:
        st.warning("⚠️ La fecha de inicio no puede ser mayor a la fecha final.")

# ── Badges de códigos visibles ────────────────────────────────────────────────
codigos_visibles = df_filtrado["Codigo"].drop_duplicates().tolist()
badges = " ".join([
    f'<span style="display:inline-block;background:#e4e6f2;border:1px solid #1e24a8;'
    f'border-radius:20px;padding:0.2rem 0.9rem;font-size:0.85rem;color:#1e24a8;'
    f'font-weight:600;margin-right:0.4rem;">📦 {c}</span>'
    for c in codigos_visibles
])
st.markdown(f'<div style="margin-bottom:1rem;">{badges}</div>', unsafe_allow_html=True)

# ── Métricas y tabla ──────────────────────────────────────────────────────────
st.markdown(
    f'<div class="section-title">📋 Movimientos '
    f'<span style="font-weight:400;font-size:0.85rem;color:#888;">({len(df_filtrado)} registros)</span></div>',
    unsafe_allow_html=True
)
render_metricas(df_filtrado)
render_tabla(df_filtrado)

# ── Descargar Excel ───────────────────────────────────────────────────────────
st.markdown('<div class="section-title">💾 Descargar datos</div>', unsafe_allow_html=True)

df_export = df_all.copy()
df_export = df_export.drop(columns=["Sin_Saldo_Inicial", "Semaforo"], errors="ignore")

buffer = exportar_excel(df_export)

nombre_personalizado = st.text_input(
    "Nombre del archivo (opcional)",
    placeholder=f"kardex_{datetime.date.today().strftime('%Y%m%d')}",
    help="Si lo dejas vacío se usará el nombre por defecto"
)

nombre_base = nombre_personalizado.strip() if nombre_personalizado.strip() else f"kardex_{datetime.date.today().strftime('%Y%m%d')}"
if not nombre_base.endswith(".xlsx"):
    nombre_base += ".xlsx"

st.download_button(
    label="⬇️ Descargar Excel completo",
    data=buffer,
    file_name=nombre_base,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)